#tany.py
from openpyxl import *

FILE='tanya1.xlsx'
BOOKS_NUM=8
SAVE=False

wb = load_workbook(FILE)
ws = wb.worksheets[2] #выбираем 3 лист
initial_values=[ws["B{}".format(i)].value for i in range(3,11)] #значение книг по умолчанию
print(initial_values)
books=[ws["A{}".format(i)].value for i in range(3,11)]
ws = wb.worksheets[3] #4 лист
coll="B"
def inc_coll(coll):
    l=len(coll)
    if l==1:
        return chr(ord(coll)+1)
    elif l==2:
        if coll=="zz":
            return "aaa"
        if coll[1]=="z":
            return chr(ord(coll[0])+1)+"a"
        return coll[0]+chr(ord(coll[1])+1)
while coll!="DA":
    cell_value=ws[coll+"15"].value
    """
    if cell_value==0:
        print('continue')
        continue
    """
    if cell_value==1:
        cell_value = ws[coll+"16"].value
        if cell_value in books:
            #в ячейке книга, смотрим ее индекс и убавляем его
            if not SAVE:
                SAVE=True
            index = books.index(cell_value)
            initial_values[index]-=1
            coll=inc_coll(coll)
    else:
        cell_values = ws[coll+"16"].value,ws[coll+"17"].value
        for i in cell_values:
            if i in books:
                index = books.index(cell_value)
                initial_values[index]-=1
        coll=inc_coll(coll)

assert len(initial_values)==BOOKS_NUM
#изменяем при необходимости книги

print(initial_values)
exit()
ws = wb.worksheets[2] #выбираем 3 лист
for i in range(BOOKS_NUM):
    ws["B{}".format(i+2)].value=initial_values[i] #значение B(i+2) равно initial_values[i]
if SAVE:
    wb.save(FILE) #сохраняем файл
